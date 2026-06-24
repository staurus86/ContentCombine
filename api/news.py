"""
News and editorial API handlers extracted from web.py.
Each function is standalone (no self), returns a dict for JSON serialization.
"""

import json
import logging
import threading
import time as _time

from storage.database import get_connection, _is_postgres, update_news_status

logger = logging.getLogger(__name__)

_JSON_FIELDS = ['bigrams', 'trigrams', 'trends_data', 'keyso_data', 'viral_data', 'tags_data', 'entity_names', 'score_breakdown']

# Telegram-источники изолированы в собственную вкладку и не должны протекать
# в ленту, кейсы, сюжеты, дайджесты и виральность. Опознаются по префиксу "TG:"
# в имени источника (та же конвенция, что в is_case_content).
# ВАЖНО: паттерн передаётся ПАРАМЕТРОМ, не литералом в SQL. На Postgres psycopg2
# %-форматирует строку запроса, когда execute идёт с params, и литеральный '%'
# в 'TG:%' роняет запрос (см. память psycopg2-literal-percent-crash).
TG_LIKE_PATTERN = "TG:%"


def _parse_json_fields(row: dict) -> dict:
    """Parse JSON string fields into Python objects for cleaner API responses."""
    for field in _JSON_FIELDS:
        val = row.get(field)
        if isinstance(val, str) and val:
            try:
                row[field] = json.loads(val)
            except (json.JSONDecodeError, ValueError):
                pass
    return row


# ---------------------------------------------------------------------------
# GET endpoints
# ---------------------------------------------------------------------------

def get_news_unified(query_params):
    """Unified news endpoint. Use ?view=editorial|final|all (default: editorial).
    Replaces get_news, get_editorial, get_final."""
    conn = get_connection()
    cur = conn.cursor()
    try:
        qs = query_params
        view = qs.get("view", ["editorial"])[0]
        limit = int(qs.get("limit", [100])[0])
        offset = int(qs.get("offset", [0])[0])
        status_filter = qs.get("status", [None])[0]
        source_filter = qs.get("source", [None])[0]
        date_from = qs.get("date_from", [None])[0]
        date_to = qs.get("date_to", [None])[0]
        min_score = int(qs.get("min_score", [0])[0])
        max_score = int(qs.get("max_score", [0])[0])
        score_filter = qs.get("score_filter", [None])[0]
        viral_level = qs.get("viral_level", [None])[0]
        tier_filter = qs.get("tier", [None])[0]
        search = qs.get("q", [None])[0]
        llm_filter = qs.get("llm", [None])[0]
        sort_field = qs.get("sort", ["parsed_at"])[0]
        sort_dir = qs.get("dir", ["desc"])[0]
        include_deleted = qs.get("deleted", ["0"])[0] == "1"
        cases_only = qs.get("cases", ["0"])[0] == "1"
        tg_mode = qs.get("tg", ["exclude"])[0]  # exclude (default) | only

        _ph = "%s" if _is_postgres() else "?"
        conditions = []
        params = []
        join_type = "LEFT JOIN"

        # Soft-delete filter
        if not include_deleted:
            conditions.append("COALESCE(n.is_deleted, 0) = 0")

        # Telegram routing: by default keep TG out of every shared view; the
        # dedicated Telegram tab passes tg=only to get just those sources.
        # Pattern is a bound parameter — never a literal (psycopg2 % safety).
        if tg_mode == "only":
            conditions.append(f"n.source LIKE {_ph}")
        else:
            conditions.append(f"n.source NOT LIKE {_ph}")
        params.append(TG_LIKE_PATTERN)

        # Cases tab: only bookmarked / research items
        if cases_only:
            conditions.append("COALESCE(n.is_case, 0) = 1")

        # View-specific default status filters
        if view == "final":
            join_type = "JOIN"
            if not status_filter:
                conditions.append("n.status IN ('processed', 'ready')")
                conditions.append("LOWER(a.llm_recommendation) = 'publish_now'")
            sort_field = qs.get("sort", ["total_score"])[0]
        elif view == "all":
            pass  # no default status filter
        else:  # editorial (default)
            if status_filter:
                conditions.append(f"n.status = {_ph}")
                params.append(status_filter)
            else:
                conditions.append("n.status NOT IN ('duplicate', 'rejected')")

        # Common filters
        if view == "final" and status_filter:
            conditions.append(f"n.status = {_ph}")
            params.append(status_filter)
        if source_filter:
            conditions.append(f"n.source = {_ph}")
            params.append(source_filter)
        if date_from:
            conditions.append(f"COALESCE(n.published_ts, n.parsed_at) >= {_ph}")
            params.append(date_from)
        if date_to:
            conditions.append(f"COALESCE(n.published_ts, n.parsed_at) <= {_ph}")
            params.append(date_to + "T23:59:59")
        if min_score > 0:
            conditions.append(f"COALESCE(a.total_score, 0) >= {_ph}")
            params.append(min_score)
        if max_score > 0:
            conditions.append(f"COALESCE(a.total_score, 0) <= {_ph}")
            params.append(max_score)
        if score_filter == "zero":
            conditions.append("COALESCE(a.total_score, 0) = 0")
        elif score_filter == "nonzero":
            conditions.append("COALESCE(a.total_score, 0) > 0")
        if viral_level:
            conditions.append(f"a.viral_level = {_ph}")
            params.append(viral_level)
        if tier_filter:
            conditions.append(f"a.entity_best_tier = {_ph}")
            params.append(tier_filter)
        if search:
            conditions.append(f"LOWER(n.title) LIKE {_ph}")
            params.append(f"%{search.lower()}%")
        if llm_filter:
            if llm_filter == "has_rec":
                conditions.append("a.llm_recommendation IS NOT NULL AND a.llm_recommendation != ''")
            elif llm_filter == "no_rec":
                conditions.append("(a.llm_recommendation IS NULL OR a.llm_recommendation = '')")
            else:
                conditions.append(f"LOWER(a.llm_recommendation) LIKE {_ph}")
                params.append(f"%{llm_filter.lower()}%")

        where = "WHERE " + " AND ".join(conditions) if conditions else ""

        # Count
        cur.execute(f"SELECT COUNT(*) FROM news n {join_type} news_analysis a ON n.id = a.news_id {where}", params[:])
        total_count = cur.fetchone()[0]

        # Status stats (for editorial view)
        status_counts = {}
        if view == "editorial":
            cur.execute(f"SELECT status, COUNT(*) FROM news WHERE COALESCE(is_deleted, 0) = 0 AND source NOT LIKE {_ph} GROUP BY status", (TG_LIKE_PATTERN,))
            for row in cur.fetchall():
                status_counts[row[0]] = row[1]

        # Sort
        allowed_sorts = {
            "total_score": "COALESCE(a.total_score, 0)",
            "final_score": "COALESCE(a.total_score, 0)",
            "viral_score": "COALESCE(a.viral_score, 0)",
            "freshness_hours": "COALESCE(a.freshness_hours, -1)",
            "source": "n.source",
            "parsed_at": "COALESCE(n.published_ts, n.parsed_at)",
        }
        order_col = allowed_sorts.get(sort_field, "n.parsed_at")
        order_dir = "ASC" if sort_dir == "asc" else "DESC"

        query = f"""
            SELECT n.id, n.source, n.title, n.description, n.url, n.h1,
                   n.published_at, n.parsed_at, n.status,
                   COALESCE(n.word_count, 0) as word_count,
                   COALESCE(n.image_count, 0) as image_count,
                   COALESCE(n.is_case, 0) as is_case,
                   COALESCE(n.published_ts, n.parsed_at) as published_ts,
                   COALESCE(a.total_score, 0) as total_score,
                   COALESCE(a.quality_score, 0) as quality_score,
                   COALESCE(a.relevance_score, 0) as relevance_score,
                   COALESCE(a.viral_score, 0) as viral_score,
                   COALESCE(a.viral_level, '') as viral_level,
                   COALESCE(a.viral_data, '[]') as viral_data,
                   COALESCE(a.sentiment_label, '') as sentiment_label,
                   COALESCE(a.sentiment_score, 0) as sentiment_score,
                   COALESCE(a.freshness_status, '') as freshness_status,
                   COALESCE(a.freshness_hours, -1) as freshness_hours,
                   COALESCE(a.tags_data, '[]') as tags_data,
                   COALESCE(a.momentum_score, 0) as momentum_score,
                   COALESCE(a.headline_score, 0) as headline_score,
                   COALESCE(a.all_checks_pass, 0) as all_checks_pass,
                   COALESCE(a.entity_names, '[]') as entity_names,
                   COALESCE(a.entity_best_tier, '') as entity_best_tier,
                   COALESCE(a.reviewed_at, '') as reviewed_at,
                   COALESCE(a.score_breakdown, '{{}}') as score_breakdown,
                   a.bigrams, a.trigrams, a.llm_recommendation, a.llm_trend_forecast,
                   a.keyso_data, a.trends_data, a.sheets_row, a.processed_at
            FROM news n
            {join_type} news_analysis a ON n.id = a.news_id
            {where}
            ORDER BY {order_col} {order_dir}
            LIMIT {_ph} OFFSET {_ph}
        """
        params.append(limit)
        params.append(offset)
        cur.execute(query, params)

        if _is_postgres():
            columns = [desc[0] for desc in cur.description]
            rows = [_parse_json_fields(dict(zip(columns, row))) for row in cur.fetchall()]
        else:
            rows = [_parse_json_fields(dict(row)) for row in cur.fetchall()]

        result = {"news": rows, "total": total_count, "limit": limit, "offset": offset}
        if status_counts:
            result["stats"] = status_counts
        return result
    finally:
        cur.close()


def _segment_analytics(seg_where, seg_params):
    """Shared analytics for a news segment (TG channels, cases, …): per-source
    averages, popular tags (24h / all-time), period summary, top by score in 24h.
    seg_where — SQL filter (e.g. 'n.source LIKE %s'); seg_params — its params."""
    from collections import Counter
    from datetime import datetime, timezone, timedelta
    conn = get_connection()
    cur = conn.cursor()
    cutoff = datetime.now(timezone.utc) - timedelta(hours=24)

    def _parse_dt(s):
        if not s:
            return None
        try:
            return datetime.fromisoformat(str(s).replace("Z", "+00:00"))
        except Exception:
            return None

    week_cutoff = datetime.now(timezone.utc) - timedelta(days=7)

    try:
        # Single pass over all TG posts → global summary/tags + per-author dashboard
        cur.execute(f"""
            SELECT n.source, COALESCE(a.total_score, 0), COALESCE(a.relevance_score, 0),
                   COALESCE(a.viral_score, 0), COALESCE(n.published_ts, n.parsed_at),
                   COALESCE(a.tags_data, '[]')
            FROM news n LEFT JOIN news_analysis a ON a.news_id = n.id
            WHERE {seg_where} AND COALESCE(n.is_deleted, 0) = 0
        """, seg_params)
        all_rows, day_rows = [], []
        tags_all, tags_day = Counter(), Counter()
        authors = {}
        for src, sc, rel, vir, dt, td in cur.fetchall():
            d = _parse_dt(dt) or datetime.min.replace(tzinfo=timezone.utc)
            is_day = d > cutoff
            is_week = d > week_cutoff
            all_rows.append((sc, rel, vir))
            if is_day:
                day_rows.append((sc, rel, vir))
            try:
                tags = json.loads(td) if isinstance(td, str) else (td or [])
            except Exception:
                tags = []
            au = authors.setdefault(src, {"sc": [], "rel": [], "vir": [], "week": 0, "day": 0, "tags": Counter()})
            au["sc"].append(sc); au["rel"].append(rel); au["vir"].append(vir)
            if is_week:
                au["week"] += 1
            if is_day:
                au["day"] += 1
            for t in tags or []:
                lbl = (t.get("label") or t.get("id") or "").strip() if isinstance(t, dict) else str(t).strip()
                if len(lbl) >= 2:
                    tags_all[lbl] += 1
                    au["tags"][lbl] += 1
                    if is_day:
                        tags_day[lbl] += 1

        def _avg(xs):
            return round(sum(xs) / len(xs), 1) if xs else 0
        by_author = [{
            "source": src,
            "avg_score": _avg(a["sc"]),
            "avg_relevance": _avg(a["rel"]),
            "avg_viral": _avg(a["vir"]),
            "count": len(a["sc"]),
            "posts_week": a["week"],
            "posts_day": a["day"],
            "top_tags": [l for l, _ in a["tags"].most_common(2)],
        } for src, a in authors.items()]
        by_author.sort(key=lambda x: (-x["posts_week"], -x["avg_score"]))

        def _summary(rows):
            n = len(rows)
            if not n:
                return {"posts": 0, "avg_score": 0, "avg_relevance": 0, "avg_viral": 0}
            return {
                "posts": n,
                "avg_score": round(sum(r[0] for r in rows) / n, 1),
                "avg_relevance": round(sum(r[1] for r in rows) / n, 1),
                "avg_viral": round(sum(r[2] for r in rows) / n, 1),
            }

        # Top posts by score in the last 24h
        if _is_postgres():
            cur.execute(f"""
                SELECT n.source, n.title, n.url, COALESCE(a.total_score, 0),
                       COALESCE(n.published_ts, n.parsed_at)
                FROM news n LEFT JOIN news_analysis a ON a.news_id = n.id
                WHERE {seg_where} AND COALESCE(n.is_deleted, 0) = 0
                  AND COALESCE(n.published_ts, n.parsed_at)::timestamptz > (NOW() - INTERVAL '24 hours')
                ORDER BY COALESCE(a.total_score, 0) DESC LIMIT 10
            """, seg_params)
        else:
            cur.execute(f"""
                SELECT n.source, n.title, n.url, COALESCE(a.total_score, 0),
                       COALESCE(n.published_ts, n.parsed_at)
                FROM news n LEFT JOIN news_analysis a ON a.news_id = n.id
                WHERE {seg_where} AND COALESCE(n.is_deleted, 0) = 0
                  AND COALESCE(n.published_ts, n.parsed_at) > datetime('now', '-1 day')
                ORDER BY COALESCE(a.total_score, 0) DESC LIMIT 10
            """, seg_params)
        top_score_24h = [{
            "source": r[0], "title": r[1], "url": r[2], "total_score": r[3], "date": str(r[4] or "")[:16],
        } for r in cur.fetchall()]

        return {
            "by_author": by_author,
            "summary": {"day": _summary(day_rows), "all": _summary(all_rows)},
            "tags_24h": [{"label": l, "count": c} for l, c in tags_day.most_common(10)],
            "tags_all": [{"label": l, "count": c} for l, c in tags_all.most_common(10)],
            "top_score_24h": top_score_24h,
        }
    finally:
        cur.close()


# Subscriber counts per channel — cached 12h (counts change slowly).
_tg_subs_cache = {}  # channel handle -> (count, ts)


def _ensure_subs_table(cur):
    cur.execute("""CREATE TABLE IF NOT EXISTS tg_subs_history (
        channel TEXT, snapshot_date TEXT, subscribers INTEGER,
        PRIMARY KEY (channel, snapshot_date))""")


def _store_subs_snapshot(counts):
    """Upsert today's subscriber count per channel (one row per channel per day)."""
    from datetime import datetime, timezone
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    conn = get_connection()
    cur = conn.cursor()
    try:
        _ensure_subs_table(cur)
        for ch, cnt in counts.items():
            if cnt is None:
                continue
            if _is_postgres():
                cur.execute("""INSERT INTO tg_subs_history (channel, snapshot_date, subscribers)
                               VALUES (%s, %s, %s)
                               ON CONFLICT (channel, snapshot_date) DO UPDATE SET subscribers = EXCLUDED.subscribers""",
                            (ch, today, cnt))
            else:
                cur.execute("INSERT OR REPLACE INTO tg_subs_history (channel, snapshot_date, subscribers) VALUES (?, ?, ?)",
                            (ch, today, cnt))
        if not _is_postgres():
            conn.commit()
    except Exception as e:
        logger.warning("Store subs snapshot failed: %s", e)
    finally:
        cur.close()


def _get_prev_subs(channels):
    """Most recent snapshot strictly before today, per channel — for the delta."""
    from datetime import datetime, timezone
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    conn = get_connection()
    cur = conn.cursor()
    _ph = "%s" if _is_postgres() else "?"
    out = {}
    try:
        _ensure_subs_table(cur)
        cur.execute(f"""SELECT t.channel, t.subscribers FROM tg_subs_history t
            WHERE t.snapshot_date = (SELECT MAX(t2.snapshot_date) FROM tg_subs_history t2
                                     WHERE t2.channel = t.channel AND t2.snapshot_date < {_ph})""",
                    (today,))
        for ch, cnt in cur.fetchall():
            out[ch] = cnt
    except Exception as e:
        logger.warning("Get prev subs failed: %s", e)
    finally:
        cur.close()
    return out


def _tg_subscribers(channels, force=False):
    """Fetch subscriber counts via the bot (cached 12h, parallel). Stores a daily
    snapshot for freshly fetched channels. Returns {handle: count|None}."""
    import time
    import config
    token = getattr(config, "TELEGRAM_BOT_TOKEN", "")
    handles = [c for c in set(channels) if c]
    if not token or not handles:
        return {c: _tg_subs_cache.get(c, (None, 0))[0] for c in channels}
    now = time.time()
    need = handles if force else [c for c in handles if now - _tg_subs_cache.get(c, (0, 0))[1] > 12 * 3600]

    def _fetch(ch):
        try:
            import requests
            r = requests.get(f"https://api.telegram.org/bot{token}/getChatMemberCount",
                             params={"chat_id": "@" + ch}, timeout=10).json()
            return ch, (r.get("result") if r.get("ok") else None)
        except Exception:
            return ch, None

    if need:
        from concurrent.futures import ThreadPoolExecutor
        fetched = {}
        try:
            with ThreadPoolExecutor(max_workers=10) as ex:
                for ch, cnt in ex.map(_fetch, need):
                    if cnt is not None:
                        _tg_subs_cache[ch] = (cnt, now)
                        fetched[ch] = cnt
        except Exception as e:
            logger.warning("TG subscriber fetch failed: %s", e)
        if fetched:
            _store_subs_snapshot(fetched)
    return {c: _tg_subs_cache.get(c, (None, 0))[0] for c in channels}


def refresh_tg_subscribers():
    """Daily job: force-refresh all TG channel subscriber counts and snapshot them."""
    import config
    chans = [(s.get("channel") or "").lstrip("@") for s in config.SOURCES if s.get("type") == "telegram"]
    _tg_subscribers([c for c in chans if c], force=True)
    logger.info("TG subscribers snapshot refreshed (%d channels)", len([c for c in chans if c]))


def get_telegram_analytics():
    """Analytics for the Telegram tab (only TG: sources), enriched with subscribers + daily delta."""
    _ph = "%s" if _is_postgres() else "?"
    res = _segment_analytics(f"n.source LIKE {_ph}", ("TG:%",))
    try:
        import config
        name_to_chan = {s["name"]: (s.get("channel") or "").lstrip("@")
                        for s in config.SOURCES if s.get("type") == "telegram"}
        chans = [name_to_chan.get(a["source"]) for a in res.get("by_author", [])]
        subs = _tg_subscribers(chans)
        prev = _get_prev_subs([c for c in chans if c])
        for a in res.get("by_author", []):
            ch = name_to_chan.get(a["source"])
            cur_cnt = subs.get(ch) if ch else None
            a["subscribers"] = cur_cnt
            p = prev.get(ch) if ch else None
            a["subs_delta"] = (cur_cnt - p) if (cur_cnt is not None and p is not None) else None
    except Exception as e:
        logger.warning("Subscriber enrich failed: %s", e)
    return res


def get_cases_analytics():
    """Analytics for the Cases tab (only is_case=1 items)."""
    return _segment_analytics("COALESCE(n.is_case, 0) = 1", ())


def get_tg_channels_digest(period="day", send=False):
    """Generate a digest of Telegram channels for a period.

    period: day (24h) | prev_day (24-48h ago) | week (7d) | month (30d).
    Pulls fresh TG posts ranked by score+virality and asks the LLM for an
    engaging, subscriber-friendly digest with links to the posts. send=False
    (default) only generates and saves — it never posts to the channel.
    """
    PERIODS = {
        "day":      {"label": "за сутки",         "limit": 7,  "from_h": 24,      "to_h": 0},
        "prev_day": {"label": "за прошлые сутки",  "limit": 7,  "from_h": 48,      "to_h": 24},
        "week":     {"label": "за неделю",         "limit": 10, "from_h": 24 * 7,  "to_h": 0},
        "month":    {"label": "за месяц",          "limit": 12, "from_h": 24 * 30, "to_h": 0},
    }
    cfg = PERIODS.get(period, PERIODS["day"])
    period_label, item_limit, from_h, to_h = cfg["label"], cfg["limit"], cfg["from_h"], cfg["to_h"]

    conn = get_connection()
    cur = conn.cursor()
    _ph = "%s" if _is_postgres() else "?"
    TG = "TG:%"
    try:
        if _is_postgres():
            time_cond = f"COALESCE(n.published_ts, n.parsed_at)::timestamptz > (NOW() - INTERVAL '{from_h} hours')"
            if to_h:
                time_cond += f" AND COALESCE(n.published_ts, n.parsed_at)::timestamptz <= (NOW() - INTERVAL '{to_h} hours')"
        else:
            time_cond = f"COALESCE(n.published_ts, n.parsed_at) > datetime('now', '-{from_h} hours')"
            if to_h:
                time_cond += f" AND COALESCE(n.published_ts, n.parsed_at) <= datetime('now', '-{to_h} hours')"
        cur.execute(f"""
            SELECT n.id, n.title, n.source, n.url, n.published_at,
                   COALESCE(a.total_score, 0), COALESCE(a.viral_score, 0),
                   COALESCE(a.viral_level, ''), COALESCE(a.tags_data, '[]')
            FROM news n LEFT JOIN news_analysis a ON a.news_id = n.id
            WHERE n.source LIKE {_ph} AND COALESCE(n.is_deleted, 0) = 0
              AND {time_cond}
            ORDER BY (COALESCE(a.total_score, 0) + COALESCE(a.viral_score, 0)) DESC
            LIMIT 40
        """, (TG,))
        cols = ["id", "title", "source", "url", "published_at", "total_score",
                "viral_score", "viral_level", "tags_data"]
        news_list = [dict(zip(cols, row)) for row in cur.fetchall()]
    finally:
        cur.close()

    from apis.digest import generate_tg_channels_digest
    result = generate_tg_channels_digest(news_list, period_label=period_label, limit=item_limit)

    # Save to digests list (style tg_<period>) for history
    try:
        import uuid
        from datetime import datetime, timezone
        from storage.database import save_digest
        if result.get("news_count", 0) > 0:
            save_digest(
                digest_id=str(uuid.uuid4())[:12],
                digest_date=datetime.now(timezone.utc).strftime("%Y-%m-%d"),
                style="tg_" + period,
                title=result.get("title", ""),
                text=result.get("text", ""),
                news_count=result.get("news_count", 0),
            )
    except Exception as e:
        logger.warning("TG digest save failed: %s", e)

    return {"status": "ok", "digest": result, "period": period}


def export_news_xlsx(query_params) -> bytes:
    """Export current filtered view as XLSX file. Returns bytes of the workbook."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # Remove limit cap — export ALL matching rows
    query_params["limit"] = ["5000"]
    query_params["offset"] = ["0"]
    result = get_news_unified(query_params)
    rows = result.get("news", [])

    wb = Workbook()
    ws = wb.active
    ws.title = "ContentCombine Export"

    # Headers (free-tier: no status / no LLM columns)
    headers = [
        "ID", "Источник", "Заголовок", "Скор", "Качество", "Релевантность",
        "Вирал", "Вирал уровень", "Свежесть (ч)", "Тон", "Тир",
        "Headline", "Momentum", "Дата парсинга", "Дата публикации",
        "Теги", "Entities", "URL"
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1DA1F2", end_color="1DA1F2", fill_type="solid")
    thin_border = Border(
        bottom=Side(style="thin", color="CCCCCC")
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, n in enumerate(rows, 2):
        tags = n.get("tags_data", [])
        if isinstance(tags, str):
            try:
                tags = json.loads(tags)
            except Exception:
                tags = []
        tags_str = ", ".join(
            t.get("label", t.get("id", "")) if isinstance(t, dict) else str(t)
            for t in tags
        ) if tags else ""

        entities = n.get("entity_names", [])
        if isinstance(entities, str):
            try:
                entities = json.loads(entities)
            except Exception:
                entities = []
        ent_str = ", ".join(entities) if entities else ""

        values = [
            n.get("id", ""),
            n.get("source", ""),
            n.get("title", ""),
            n.get("total_score", 0),
            n.get("quality_score", 0),
            n.get("relevance_score", 0),
            n.get("viral_score", 0),
            n.get("viral_level", ""),
            n.get("freshness_hours", ""),
            n.get("sentiment_label", ""),
            n.get("entity_best_tier", ""),
            n.get("headline_score", 0),
            n.get("momentum_score", 0),
            (n.get("parsed_at") or "")[:16].replace("T", " "),
            (n.get("published_at") or "")[:16].replace("T", " "),
            tags_str,
            ent_str,
            n.get("url", ""),
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

        # Color-code score
        score = n.get("total_score", 0) or 0
        score_cell = ws.cell(row=row_idx, column=4)
        if score >= 70:
            score_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif score >= 40:
            score_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        elif score > 0:
            score_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Column widths
    col_widths = [6, 15, 50, 8, 8, 8, 8, 10, 10, 10, 6, 8, 8, 16, 16, 30, 25, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# Deprecated: use get_news_unified(view=...) instead
def get_news(query_params):
    query_params.setdefault("view", ["all"])
    if "status" not in query_params:
        query_params["status"] = [None]
    # Preserve old default: only approved/processed/ready
    qs = dict(query_params)
    if not qs.get("status", [None])[0]:
        qs["status"] = [""]
    qs["view"] = ["all"]
    # Old behavior: filter approved/processed/ready by default
    conn = get_connection()
    cur = conn.cursor()
    try:
        _qs = query_params
        limit = int(_qs.get("limit", [100])[0])
        offset = int(_qs.get("offset", [0])[0])
        status_filter = _qs.get("status", [None])[0]
        source_filter = _qs.get("source", [None])[0]
        date_from = _qs.get("date_from", [None])[0]
        date_to = _qs.get("date_to", [None])[0]
        llm_filter = _qs.get("llm", [None])[0]
        _ph = "%s" if _is_postgres() else "?"
        conditions = ["COALESCE(n.is_deleted, 0) = 0"]
        params = []
        if status_filter:
            conditions.append(f"n.status = {_ph}")
            params.append(status_filter)
        else:
            conditions.append("n.status IN ('approved', 'processed', 'ready')")
        if source_filter:
            conditions.append(f"n.source = {_ph}")
            params.append(source_filter)
        if date_from:
            conditions.append(f"COALESCE(n.published_ts, n.parsed_at) >= {_ph}")
            params.append(date_from)
        if date_to:
            conditions.append(f"COALESCE(n.published_ts, n.parsed_at) <= {_ph}")
            params.append(date_to + "T23:59:59")
        if llm_filter:
            if llm_filter == "has_rec":
                conditions.append("a.llm_recommendation IS NOT NULL AND a.llm_recommendation != ''")
            elif llm_filter == "no_rec":
                conditions.append("(a.llm_recommendation IS NULL OR a.llm_recommendation = '')")
            else:
                conditions.append(f"LOWER(a.llm_recommendation) LIKE {_ph}")
                params.append(f"%{llm_filter.lower()}%")
        where = "WHERE " + " AND ".join(conditions) if conditions else ""
        cur.execute(f"SELECT COUNT(*) FROM news n LEFT JOIN news_analysis a ON n.id = a.news_id {where}", params[:])
        total_count = cur.fetchone()[0]
        query = f"""
            SELECT n.id, n.source, n.title, n.url, n.h1, n.description,
                   n.published_at, n.parsed_at, n.status,
                   a.bigrams, a.trigrams, a.trends_data, a.keyso_data,
                   a.llm_recommendation, a.llm_trend_forecast, a.sheets_row, a.processed_at,
                   a.viral_score, a.viral_level, a.viral_data,
                   a.sentiment_label, a.sentiment_score,
                   a.freshness_status, a.freshness_hours,
                   a.tags_data, a.momentum_score, a.headline_score, a.total_score
            FROM news n LEFT JOIN news_analysis a ON n.id = a.news_id
            {where} ORDER BY n.parsed_at DESC LIMIT {_ph} OFFSET {_ph}
        """
        params.append(limit)
        params.append(offset)
        cur.execute(query, params)
        if _is_postgres():
            columns = [desc[0] for desc in cur.description]
            rows = [_parse_json_fields(dict(zip(columns, row))) for row in cur.fetchall()]
        else:
            rows = [_parse_json_fields(dict(row)) for row in cur.fetchall()]
        return {"news": rows, "total": total_count, "limit": limit, "offset": offset}
    finally:
        cur.close()


def get_editorial(query_params):
    """Deprecated: proxies to get_news_unified(view=editorial)."""
    query_params.setdefault("view", ["editorial"])
    return get_news_unified(query_params)


def get_final(query_params):
    """Deprecated: proxies to get_news_unified(view=final)."""
    query_params.setdefault("view", ["final"])
    return get_news_unified(query_params)


def get_moderation_list(query_params):
    """Returns news with status 'moderation' plus local analysis data."""
    limit = int(query_params.get("limit", ["100"])[0])
    offset = int(query_params.get("offset", ["0"])[0])
    source = query_params.get("source", [""])[0]
    min_score = int(query_params.get("min_score", ["0"])[0])
    q = query_params.get("q", [""])[0]

    conn = get_connection()
    cur = conn.cursor()
    ph = "%s" if _is_postgres() else "?"

    conditions = ["n.status = 'moderation'", "COALESCE(n.is_deleted, 0) = 0"]
    params = []

    if source:
        conditions.append(f"n.source = {ph}")
        params.append(source)
    if min_score > 0:
        conditions.append(f"COALESCE(na.total_score, 0) >= {ph}")
        params.append(min_score)
    if q:
        conditions.append(f"LOWER(n.title) LIKE {ph}")
        params.append(f"%{q.lower()}%")

    where = " AND ".join(conditions)

    try:
        cur.execute(f"""
            SELECT n.id, n.source, n.title, n.url, n.published_at, n.parsed_at, n.status,
                   n.description,
                   na.total_score, na.quality_score, na.relevance_score,
                   na.freshness_hours, na.viral_score, na.viral_data,
                   na.sentiment_label,
                   na.tags_data as tags, na.entity_names as entities, na.headline_score, na.momentum_score
            FROM news n
            LEFT JOIN news_analysis na ON na.news_id = n.id
            WHERE {where}
            ORDER BY n.parsed_at DESC
            LIMIT {ph} OFFSET {ph}
        """, (*params, limit, offset))

        if _is_postgres():
            columns = [d[0] for d in cur.description]
            rows = [dict(zip(columns, r)) for r in cur.fetchall()]
        else:
            rows = [dict(r) for r in cur.fetchall()]

        cur.execute(f"SELECT COUNT(*), COALESCE(AVG(na.total_score), 0) FROM news n LEFT JOIN news_analysis na ON na.news_id = n.id WHERE {where}", tuple(params))
        row = cur.fetchone()
        total = row[0]
        avg_score = round(row[1], 1) if row[1] else 0
    finally:
        cur.close()

    return {"status": "ok", "news": rows, "total": total, "avg_score": avg_score}


def get_moderation(body):
    """POST version for compatibility — delegates to get_moderation_list."""
    return get_moderation_list({})


def get_event_chain_by_id(news_id):
    """Returns event chain for a given news (GET endpoint)."""
    try:
        conn = get_connection()
        cur = conn.cursor()
        ph = "%s" if _is_postgres() else "?"
        try:
            cur.execute(f"SELECT id, source, title, published_at, status FROM news WHERE id = {ph}", (news_id,))
            row = cur.fetchone()
            if not row:
                return {"status": "error", "message": "news not found"}
            if _is_postgres():
                columns = [desc[0] for desc in cur.description]
                news = dict(zip(columns, row))
            else:
                news = dict(row)
        finally:
            cur.close()
        from checks.temporal_clusters import get_event_chain
        return get_event_chain(news)
    except Exception as e:
        logger.error(f"Event chain error: {e}")
        return {"status": "error", "message": str(e), "chain": [], "chain_length": 0, "days_span": 0, "phase": "single"}


def get_event_chain(body):
    """POST version of event chain lookup."""
    news_id = body.get("news_id", "")
    if not news_id:
        return {"status": "error", "message": "news_id required"}
    conn = get_connection()
    cur = conn.cursor()
    try:
        ph = "%s" if _is_postgres() else "?"
        cur.execute(f"SELECT * FROM news WHERE id = {ph}", (news_id,))
        if _is_postgres():
            columns = [desc[0] for desc in cur.description]
            row = cur.fetchone()
            if not row:
                return {"status": "error", "message": "not found"}
            news = dict(zip(columns, row))
        else:
            row = cur.fetchone()
            if not row:
                return {"status": "error", "message": "not found"}
            news = dict(row)
        from checks.temporal_clusters import get_event_chain as _get_chain
        chain = _get_chain(news)
        return {"status": "ok", **chain}
    finally:
        cur.close()


# ---------------------------------------------------------------------------
# POST endpoints — review / approval / rejection
# ---------------------------------------------------------------------------

def run_review(body):
    """Run review pipeline for selected news."""
    news_ids = body.get("news_ids", [])
    if not news_ids:
        return {"status": "error", "message": "No news selected"}
    try:
        conn = get_connection()
        cur = conn.cursor()
        try:
            ph = "%s" if _is_postgres() else "?"
            placeholders = ",".join([ph] * len(news_ids))
            cur.execute(f"SELECT id, source, url, title, h1, description, plain_text, published_at, parsed_at, status FROM news WHERE id IN ({placeholders})", news_ids)
            if _is_postgres():
                columns = [desc[0] for desc in cur.description]
                news_list = [dict(zip(columns, row)) for row in cur.fetchall()]
            else:
                news_list = [dict(row) for row in cur.fetchall()]

            from checks.pipeline import run_review_pipeline
            result = run_review_pipeline(news_list)
            return {"status": "ok", **result}
        finally:
            cur.close()
    except Exception as e:
        return {"status": "error", "message": str(e), "type": type(e).__name__}


def review_batch(body):
    """Review news by status (batch, without changing status)."""
    status = body.get("status", "new")
    limit = int(body.get("limit", 50))
    try:
        conn = get_connection()
        cur = conn.cursor()
        try:
            ph = "%s" if _is_postgres() else "?"
            if status:
                cur.execute(f"SELECT id, source, url, title, h1, description, plain_text, published_at, parsed_at, status FROM news WHERE status = {ph} ORDER BY parsed_at DESC LIMIT {ph}", (status, limit))
            else:
                cur.execute(f"SELECT id, source, url, title, h1, description, plain_text, published_at, parsed_at, status FROM news ORDER BY parsed_at DESC LIMIT {ph}", (limit,))
            if _is_postgres():
                columns = [desc[0] for desc in cur.description]
                news_list = [dict(zip(columns, row)) for row in cur.fetchall()]
            else:
                news_list = [dict(row) for row in cur.fetchall()]

            if not news_list:
                return {"status": "ok", "results": [], "groups": []}

            from checks.pipeline import run_review_pipeline
            from checks.deduplication import tfidf_similarity, build_groups
            from checks.quality import check_quality
            from checks.relevance import check_relevance
            from checks.freshness import check_freshness
            from checks.viral_score import viral_score
            from checks.tags import auto_tag
            from checks.sentiment import analyze_sentiment
            from checks.momentum import get_momentum

            results = []
            for news in news_list:
                result = {
                    "id": news["id"],
                    "title": news.get("title", ""),
                    "source": news.get("source", ""),
                    "url": news.get("url", ""),
                    "published_at": news.get("published_at", ""),
                    "status": news.get("status", ""),
                    "checks": {},
                }
                result["checks"]["quality"] = check_quality(news)
                result["checks"]["relevance"] = check_relevance(news)
                result["checks"]["freshness"] = check_freshness(news)
                result["checks"]["viral"] = viral_score(news)
                result["tags"] = auto_tag(news)
                result["sentiment"] = analyze_sentiment(news)
                result["momentum"] = get_momentum(news)

                all_pass = all(c["pass"] for c in result["checks"].values())
                total_score = sum(c["score"] for c in result["checks"].values()) // 4
                momentum_bonus = result["momentum"]["score"] // 5
                total_score = min(100, total_score + momentum_bonus)
                result["overall_pass"] = all_pass
                result["total_score"] = total_score
                results.append(result)

            titles = [r["title"] for r in results]
            pairs = tfidf_similarity(titles)
            groups = build_groups(results, pairs)
            for group in groups:
                for idx in group.get("duplicate_indices", []):
                    if idx < len(results):
                        results[idx]["overall_pass"] = False
                        results[idx]["is_duplicate"] = True
                for member in group["members"]:
                    member["dedup_status"] = group["status"]

            return {"status": "ok", "results": results, "groups": groups}
        finally:
            cur.close()
    except Exception as e:
        return {"status": "error", "message": str(e), "type": type(e).__name__}


def run_auto_review(body):
    """Run auto-review in batches of 20 (saves results to DB)."""
    try:
        conn = get_connection()
        cur = conn.cursor()
        try:
            ph = "%s" if _is_postgres() else "?"
            BATCH_SIZE = 20

            cur.execute("SELECT COUNT(*) FROM news WHERE status = 'new'")
            total_pending = cur.fetchone()[0]

            if total_pending == 0:
                return {"status": "ok", "reviewed": 0, "message": "Нет новых для проверки", "remaining": 0}

            cur.execute(f"""
                SELECT id, source, url, title, h1, description, plain_text, published_at, parsed_at, status
                FROM news WHERE status = 'new'
                ORDER BY parsed_at DESC LIMIT {ph}
            """, (BATCH_SIZE,))
            if _is_postgres():
                columns = [desc[0] for desc in cur.description]
                news_list = [dict(zip(columns, row)) for row in cur.fetchall()]
            else:
                news_list = [dict(row) for row in cur.fetchall()]

            from checks.pipeline import run_review_pipeline
            result = run_review_pipeline(news_list, update_status=True)
            reviewed = len(result.get("results", []))
            dupes = sum(1 for r in result.get("results", []) if r.get("is_duplicate"))
            rejected = sum(1 for r in result.get("results", []) if r.get("auto_rejected"))
            remaining = total_pending - reviewed
            return {
                "status": "ok",
                "reviewed": reviewed,
                "duplicates": dupes,
                "auto_rejected": rejected,
                "remaining": remaining,
                "message": f"Проверено: {reviewed}, дубликатов: {dupes}, отклонено: {rejected}" +
                           (f". Осталось: {remaining}" if remaining > 0 else "")
            }
        finally:
            cur.close()
    except Exception as e:
        return {"status": "error", "message": str(e)}


def approve_news(body):
    """Approve news and start background enrichment."""
    news_ids = body.get("news_ids", [])
    if not news_ids:
        return {"status": "error", "message": "No news selected"}
    try:
        from checks.pipeline import approve_for_enrichment
        from checks.feedback import record_decision
        approve_for_enrichment(news_ids)
        for nid in news_ids:
            try:
                record_decision(nid, "approved")
            except Exception:
                pass

        def _bg_enrich(ids):
            from scheduler import _process_single_news
            for nid in ids:
                try:
                    _process_single_news(nid)
                except Exception as e:
                    logger.warning("Background enrich failed for %s: %s", nid, e)
        threading.Thread(target=_bg_enrich, args=(list(news_ids),), daemon=True).start()

        return {"status": "ok", "approved": len(news_ids), "enriching": True}
    except Exception as e:
        return {"status": "error", "message": str(e)}


def reject_news(body):
    """Reject one or multiple news."""
    news_ids = body.get("news_ids", [])
    news_id = body.get("news_id")
    if news_id and not news_ids:
        news_ids = [news_id]
    if not news_ids:
        return {"status": "error", "message": "news_ids required"}
    try:
        from checks.feedback import record_decision
        for nid in news_ids:
            update_news_status(nid, "rejected")
            try:
                record_decision(nid, "rejected")
            except Exception:
                pass
        return {"status": "ok", "rejected": len(news_ids)}
    except Exception as e:
        return {"status": "error", "message": str(e)}


def bulk_status(body):
    """Change status for multiple news."""
    news_ids = body.get("news_ids", [])
    new_status = body.get("status", "")
    if not news_ids or not new_status:
        return {"status": "error", "message": "news_ids and status required"}
    for nid in news_ids:
        update_news_status(nid, new_status)
    return {"status": "ok", "updated": len(news_ids)}


def delete_news(body):
    """Soft-delete news (sets is_deleted=1)."""
    news_ids = body.get("news_ids", [])
    if not news_ids:
        return {"status": "error", "message": "news_ids required"}
    conn = get_connection()
    cur = conn.cursor()
    try:
        if _is_postgres():
            placeholders = ",".join(["%s"] * len(news_ids))
        else:
            placeholders = ",".join(["?"] * len(news_ids))
        from datetime import datetime, timezone
        now = datetime.now(timezone.utc).isoformat()
        _ph = "%s" if _is_postgres() else "?"
        cur.execute(f"UPDATE news SET is_deleted=1, deleted_at={_ph} WHERE id IN ({placeholders})",
                    (now, *news_ids))
        # Cascade: soft-delete related articles
        cur.execute(f"UPDATE articles SET is_deleted = 1, deleted_at = {_ph} WHERE news_id IN ({placeholders}) AND COALESCE(is_deleted, 0) = 0",
                    (now, *news_ids))
        if not _is_postgres():
            conn.commit()
        # Память модерации: ручное удаление = негативный сигнал по источнику.
        # Накапливается в feedback_stats; «аудит источников» читает долю удалений.
        try:
            from checks.feedback import record_decision
            for nid in news_ids:
                try:
                    record_decision(nid, "rejected")
                except Exception:
                    pass
        except Exception:
            pass
        return {"status": "ok", "deleted": len(news_ids)}
    finally:
        cur.close()


def restore_news(body):
    """Restore soft-deleted news."""
    news_ids = body.get("news_ids", [])
    if not news_ids:
        return {"status": "error", "message": "news_ids required"}
    conn = get_connection()
    cur = conn.cursor()
    try:
        if _is_postgres():
            placeholders = ",".join(["%s"] * len(news_ids))
        else:
            placeholders = ",".join(["?"] * len(news_ids))
        cur.execute(f"UPDATE news SET is_deleted=0, deleted_at=NULL WHERE id IN ({placeholders})",
                    tuple(news_ids))
        # Cascade: restore related articles
        cur.execute(f"UPDATE articles SET is_deleted = 0, deleted_at = NULL WHERE news_id IN ({placeholders}) AND is_deleted = 1",
                    tuple(news_ids))
        if not _is_postgres():
            conn.commit()
        return {"status": "ok", "restored": len(news_ids)}
    finally:
        cur.close()


def purge_news(body):
    """Hard-delete news permanently (admin only)."""
    news_ids = body.get("news_ids", [])
    if not news_ids:
        return {"status": "error", "message": "news_ids required"}
    conn = get_connection()
    cur = conn.cursor()
    try:
        if _is_postgres():
            placeholders = ",".join(["%s"] * len(news_ids))
        else:
            placeholders = ",".join(["?"] * len(news_ids))
        cur.execute(f"DELETE FROM news_analysis WHERE news_id IN ({placeholders})", tuple(news_ids))
        cur.execute(f"DELETE FROM news WHERE id IN ({placeholders})", tuple(news_ids))
        if not _is_postgres():
            conn.commit()
        return {"status": "ok", "purged": len(news_ids)}
    finally:
        cur.close()


def get_trash(query_params):
    """Get soft-deleted news with full analysis data (same as editorial)."""
    conn = get_connection()
    cur = conn.cursor()
    try:
        limit = int(query_params.get("limit", [100])[0])
        offset = int(query_params.get("offset", [0])[0])
        source_filter = query_params.get("source", [None])[0]
        _ph = "%s" if _is_postgres() else "?"

        conditions = ["n.is_deleted = 1"]
        params = []
        if source_filter:
            conditions.append(f"n.source = {_ph}")
            params.append(source_filter)
        where = "WHERE " + " AND ".join(conditions)

        cur.execute(f"SELECT COUNT(*) FROM news n {where}", params[:])
        total = cur.fetchone()[0]

        cur.execute(f"""
            SELECT n.id, n.source, n.title, n.description, n.url, n.published_at, n.parsed_at,
                   n.status, n.deleted_at,
                   COALESCE(a.total_score, 0) as total_score,
                   COALESCE(a.quality_score, 0) as quality_score,
                   COALESCE(a.relevance_score, 0) as relevance_score,
                   COALESCE(a.viral_score, 0) as viral_score,
                   COALESCE(a.viral_level, '') as viral_level,
                   COALESCE(a.viral_data, '[]') as viral_data,
                   COALESCE(a.sentiment_label, '') as sentiment_label,
                   COALESCE(a.freshness_hours, -1) as freshness_hours,
                   COALESCE(a.tags_data, '[]') as tags_data,
                   COALESCE(a.momentum_score, 0) as momentum_score,
                   COALESCE(a.headline_score, 0) as headline_score,
                   COALESCE(a.entity_names, '[]') as entity_names,
                   a.llm_recommendation
            FROM news n
            LEFT JOIN news_analysis a ON n.id = a.news_id
            {where}
            ORDER BY n.deleted_at DESC
            LIMIT {_ph} OFFSET {_ph}
        """, (*params, limit, offset))

        if _is_postgres():
            columns = [desc[0] for desc in cur.description]
            rows = [_parse_json_fields(dict(zip(columns, row))) for row in cur.fetchall()]
        else:
            rows = [_parse_json_fields(dict(row)) for row in cur.fetchall()]

        return {"news": rows, "total": total, "limit": limit, "offset": offset}
    finally:
        cur.close()


def export_trash_to_sheets():
    """Export all deleted news to 'Удалённые' tab in Google Sheets."""
    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT n.id, n.source, n.title, n.description, n.url, n.published_at, n.parsed_at,
                   n.status, n.deleted_at,
                   COALESCE(a.total_score, 0) as total_score,
                   COALESCE(a.quality_score, 0) as quality_score,
                   COALESCE(a.relevance_score, 0) as relevance_score,
                   COALESCE(a.viral_score, 0) as viral_score,
                   COALESCE(a.viral_data, '[]') as viral_data,
                   COALESCE(a.sentiment_label, '') as sentiment_label,
                   COALESCE(a.freshness_hours, -1) as freshness_hours,
                   COALESCE(a.tags_data, '[]') as tags_data,
                   COALESCE(a.momentum_score, 0) as momentum_score,
                   COALESCE(a.headline_score, 0) as headline_score,
                   COALESCE(a.entity_names, '[]') as entity_names
            FROM news n
            LEFT JOIN news_analysis a ON n.id = a.news_id
            WHERE n.is_deleted = 1
            ORDER BY n.deleted_at DESC
        """)
        if _is_postgres():
            columns = [desc[0] for desc in cur.description]
            rows = [dict(zip(columns, row)) for row in cur.fetchall()]
        else:
            rows = [dict(row) for row in cur.fetchall()]
    finally:
        cur.close()

    if not rows:
        return {"status": "ok", "written": 0, "message": "Корзина пуста"}

    from storage.sheets import write_deleted_batch
    result = write_deleted_batch(rows)
    return {"status": "ok", **result}


def auto_purge_old_deleted(days=30):
    """Permanently delete items in trash older than N days."""
    conn = get_connection()
    cur = conn.cursor()
    _ph = "%s" if _is_postgres() else "?"
    from datetime import datetime, timezone, timedelta
    cutoff = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()
    try:
        cur.execute(f"SELECT id FROM news WHERE is_deleted = 1 AND deleted_at < {_ph}", (cutoff,))
        if _is_postgres():
            old_ids = [row[0] for row in cur.fetchall()]
        else:
            old_ids = [row[0] for row in cur.fetchall()]
        if not old_ids:
            return 0
        placeholders = ",".join([_ph] * len(old_ids))
        cur.execute(f"DELETE FROM news_analysis WHERE news_id IN ({placeholders})", tuple(old_ids))
        cur.execute(f"DELETE FROM task_queue WHERE news_id IS NOT NULL AND news_id IN ({placeholders})", tuple(old_ids))
        cur.execute(f"DELETE FROM news WHERE id IN ({placeholders})", tuple(old_ids))
        # Also purge old deleted articles
        cur.execute(f"DELETE FROM articles WHERE is_deleted = 1 AND deleted_at < {_ph}", (cutoff,))
        if not _is_postgres():
            conn.commit()
        logger.info("Auto-purged %d old deleted news items", len(old_ids))
        return len(old_ids)
    finally:
        cur.close()


def cleanup_short_news(min_chars=100):
    """Soft-delete news with title shorter than min_chars."""
    conn = get_connection()
    cur = conn.cursor()
    _ph = "%s" if _is_postgres() else "?"
    from datetime import datetime, timezone
    now = datetime.now(timezone.utc).isoformat()
    try:
        cur.execute(f"""
            UPDATE news SET is_deleted=1, deleted_at={_ph}
            WHERE COALESCE(is_deleted, 0) = 0
              AND COALESCE(is_case, 0) = 0
              AND LENGTH(title) < {_ph}
              AND status NOT IN ('approved', 'ready')
        """, (now, min_chars))
        count = cur.rowcount
        if not _is_postgres():
            conn.commit()
        if count > 0:
            logger.info("Auto-deleted %d news with title < %d chars", count, min_chars)
        return count
    finally:
        cur.close()


def cleanup_old_news(days=7):
    """Soft-delete news older than N days (except approved/ready)."""
    conn = get_connection()
    cur = conn.cursor()
    _ph = "%s" if _is_postgres() else "?"
    from datetime import datetime, timezone, timedelta
    cutoff = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()
    now = datetime.now(timezone.utc).isoformat()
    try:
        cur.execute(f"""
            UPDATE news SET is_deleted=1, deleted_at={_ph}
            WHERE COALESCE(is_deleted, 0) = 0
              AND COALESCE(is_case, 0) = 0
              AND parsed_at < {_ph}
              AND status NOT IN ('approved', 'ready')
        """, (now, cutoff))
        count = cur.rowcount
        if not _is_postgres():
            conn.commit()
        if count > 0:
            logger.info("Auto-deleted %d news older than %d days", count, days)
        return count
    finally:
        cur.close()


def _parse_any_date(s):
    """Парсит дату из ISO или RFC822; возвращает aware datetime или None."""
    if not s:
        return None
    from datetime import datetime, timezone
    from email.utils import parsedate_to_datetime
    s = str(s).strip()
    d = None
    try:
        d = datetime.fromisoformat(s.replace("Z", "+00:00"))
    except Exception:
        try:
            d = parsedate_to_datetime(s)
        except Exception:
            return None
    if d is not None and d.tzinfo is None:
        d = d.replace(tzinfo=timezone.utc)
    return d


def soft_delete_stale_news(max_age_days=None):
    """Свежесть: новости старше max_age_days уезжают в «Удалённые» (soft-delete).

    Возраст берётся по ДАТЕ СТАТЬИ (published_at), при её отсутствии — по parsed_at
    (когда собрали). Не трогает approved/ready и уже удалённые. Это soft-delete
    (is_deleted=1 + deleted_at), не физическое удаление — восстановимо из корзины.
    """
    import config
    if max_age_days is None:
        max_age_days = getattr(config, "NEWS_MAX_AGE_DAYS", 14)
    from datetime import datetime, timezone, timedelta
    now = datetime.now(timezone.utc)
    cutoff = now - timedelta(days=max_age_days)
    conn = get_connection()
    cur = conn.cursor()
    _ph = "%s" if _is_postgres() else "?"
    try:
        cur.execute("""
            SELECT id, published_at, parsed_at FROM news
            WHERE COALESCE(is_deleted, 0) = 0
              AND COALESCE(is_case, 0) = 0
              AND status NOT IN ('approved', 'ready')
        """)
        stale = []
        for r in cur.fetchall():
            nid, pub, parsed = r[0], r[1], r[2]
            eff = _parse_any_date(pub) or _parse_any_date(parsed)
            if eff is not None and eff < cutoff:
                stale.append(nid)
        now_iso = now.isoformat()
        for i in range(0, len(stale), 200):
            chunk = stale[i:i + 200]
            placeholders = ",".join([_ph] * len(chunk))
            cur.execute(
                f"UPDATE news SET is_deleted=1, deleted_at={_ph} WHERE id IN ({placeholders})",
                (now_iso, *chunk),
            )
        if not _is_postgres():
            conn.commit()
        if stale:
            logger.info("Freshness: soft-deleted %d stale news (>%dd) to trash", len(stale), max_age_days)
        return len(stale)
    finally:
        cur.close()


_CASE_RE = __import__("re").compile(r"research|исследовани|кейс", __import__("re").IGNORECASE)


def is_case_tag(tag) -> bool:
    """True, если тег — «кейс»: id=='research' или label/slug ~ research|исследование|кейс."""
    if isinstance(tag, dict):
        tid = str(tag.get("id", ""))
        label = str(tag.get("label", "")) + " " + str(tag.get("slug", ""))
    else:
        tid = label = str(tag)
    if tid.lower() == "research":
        return True
    return bool(_CASE_RE.search(label) or _CASE_RE.search(tid))


def tags_contain_case(tags_data) -> bool:
    """tags_data (JSON-строка или список) содержит тег-кейс?"""
    if isinstance(tags_data, str):
        try:
            tags_data = json.loads(tags_data or "[]")
        except Exception:
            return False
    if not isinstance(tags_data, list):
        return False
    return any(is_case_tag(t) for t in tags_data)


# «Настоящий кейс» = статья-кейс/руководство/разбор (а не любой research-тег и не анонс).
# Обязательные слова-индикаторы в заголовке/тексте (RU+EN). Телеграм-анонсы исключаются.
_CASE_CONTENT_RE = __import__("re").compile(
    r"кейс|case stud|как мы |как я |как мне |как нам |руководств|пошагов|"
    r"чек-?лист|разбор|инструкц|гайд|how (i|we) |how to |step[ -]by[ -]step|"
    r"walkthrough|tutorial|playbook|checklist|case study|\bguide\b",
    __import__("re").IGNORECASE,
)


def is_case_content(title, text=None, source="") -> bool:
    """True только если это реальный кейс/руководство — по словам-индикаторам
    в ЗАГОЛОВКЕ (кейс обычно заявлен в тайтле; так точнее, без ложных по телу).
    Телеграм-источники (TG:*) исключены — там кейсы не пишут (только анонсы)."""
    if source and str(source).startswith("TG:"):
        return False
    # 0-словные статьи (мёртвая ссылка/404 или не извлёкся текст) — не кейсы:
    # «кейс» без тела бесполезен и ведёт на битую страницу.
    if text is not None and len((text or "").strip()) < 200:
        return False
    return bool(_CASE_CONTENT_RE.search((title or "").lower()))


def reclassify_cases():
    """Разовая переразметка is_case по контент-правилу (исправляет старую широкую
    разметку по тегу). Ставит 1, где статья — кейс/руководство и НЕ из Telegram;
    снимает 1 там, где не подходит. Возвращает (помечено, снято)."""
    conn = get_connection()
    cur = conn.cursor()
    _ph = "%s" if _is_postgres() else "?"
    to_set, to_unset = [], []
    try:
        cur.execute("SELECT id, title, COALESCE(plain_text, description, ''), source, COALESCE(is_case,0) FROM news")
        for nid, title, body, source, cur_flag in cur.fetchall():
            want = 1 if is_case_content(title, body, source) else 0
            if want == 1 and not cur_flag:
                to_set.append(nid)
            elif want == 0 and cur_flag:
                to_unset.append(nid)
        for flag, ids in ((1, to_set), (0, to_unset)):
            for i in range(0, len(ids), 200):
                chunk = ids[i:i + 200]
                ph = ",".join([_ph] * len(chunk))
                cur.execute(f"UPDATE news SET is_case={_ph} WHERE id IN ({ph})", (flag, *chunk))
        if not _is_postgres():
            conn.commit()
        logger.info("reclassify_cases: +%d set, -%d unset", len(to_set), len(to_unset))
        return len(to_set), len(to_unset)
    finally:
        cur.close()


def set_case(body):
    """Поставить/снять закладку «Кейс» (is_case) на список новостей."""
    news_ids = body.get("news_ids") or ([body["news_id"]] if body.get("news_id") else [])
    if not news_ids:
        return {"status": "error", "message": "news_ids required"}
    on = 1 if body.get("on", True) else 0
    conn = get_connection()
    cur = conn.cursor()
    _ph = "%s" if _is_postgres() else "?"
    try:
        placeholders = ",".join([_ph] * len(news_ids))
        cur.execute(f"UPDATE news SET is_case={_ph} WHERE id IN ({placeholders})",
                    (on, *news_ids))
        if not _is_postgres():
            conn.commit()
        return {"status": "ok", "is_case": on, "count": len(news_ids)}
    finally:
        cur.close()


def backfill_cases():
    """Проставить is_case=1 всем новостям с тегом-кейсом (Python-проход по tags_data)."""
    conn = get_connection()
    cur = conn.cursor()
    _ph = "%s" if _is_postgres() else "?"
    try:
        cur.execute("""
            SELECT n.id, a.tags_data FROM news n
            JOIN news_analysis a ON a.news_id = n.id
            WHERE COALESCE(n.is_case, 0) = 0 AND a.tags_data IS NOT NULL
        """)
        hit = [r[0] for r in cur.fetchall() if tags_contain_case(r[1])]
        for i in range(0, len(hit), 200):
            chunk = hit[i:i + 200]
            placeholders = ",".join([_ph] * len(chunk))
            cur.execute(f"UPDATE news SET is_case=1 WHERE id IN ({placeholders})", tuple(chunk))
        if not _is_postgres():
            conn.commit()
        if hit:
            logger.info("Cases backfill: marked %d news with research tag", len(hit))
        return len(hit)
    finally:
        cur.close()


def redate_missing_dates(limit: int = 2000):
    """Re-fetch articles без реальной даты и проставить published_at/published_ts.
    Запускается ВНУТРИ контейнера (рабочие прокси/IP) — для бэкфилла исторических дат."""
    from parsers.html_parser import _fetch_article, _is_junk_url
    from storage.database import normalize_ts
    from datetime import datetime, timezone
    conn = get_connection()
    cur = conn.cursor()
    _ph = "%s" if _is_postgres() else "?"
    cur.execute(f"""SELECT id, url FROM news
                    WHERE (published_at IS NULL OR published_at='')
                      AND COALESCE(is_deleted,0)=0 AND url LIKE 'http%'
                    LIMIT {int(limit)}""")
    rows = cur.fetchall()
    fixed = junk = 0
    now = datetime.now(timezone.utc).isoformat()
    for r in rows:
        nid, url = r[0], r[1]
        if _is_junk_url(url):
            cur.execute(f"UPDATE news SET is_deleted=1, deleted_at={_ph} WHERE id={_ph} AND COALESCE(is_case,0)=0", (now, nid))
            junk += 1
            continue
        pub = ""
        try:
            _, _, _, pub, _ = _fetch_article(url)
        except Exception:
            pub = ""
        ts = normalize_ts(pub, "") if pub else ""
        if ts:
            cur.execute(f"UPDATE news SET published_at={_ph}, published_ts={_ph} WHERE id={_ph}", (pub, ts, nid))
            fixed += 1
    if not _is_postgres():
        conn.commit()
    cur.close()
    logger.info("redate_missing_dates: scanned=%d fixed=%d junk=%d", len(rows), fixed, junk)
    return {"scanned": len(rows), "fixed": fixed, "junk": junk}


def news_detail(body):
    """Get full news + analysis detail."""
    news_id = body.get("news_id")
    if not news_id:
        return {"status": "error", "message": "news_id required"}
    conn = get_connection()
    cur = conn.cursor()
    try:
        ph = "%s" if _is_postgres() else "?"
        cur.execute(f"SELECT * FROM news WHERE id = {ph}", (news_id,))
        row = cur.fetchone()
        if not row:
            return {"status": "error", "message": "Not found"}
        if _is_postgres():
            columns = [desc[0] for desc in cur.description]
            news = dict(zip(columns, row))
        else:
            news = dict(row)
        cur.execute(f"SELECT * FROM news_analysis WHERE news_id = {ph}", (news_id,))
        arow = cur.fetchone()
        analysis = None
        if arow:
            if _is_postgres():
                columns = [desc[0] for desc in cur.description]
                analysis = _parse_json_fields(dict(zip(columns, arow)))
            else:
                analysis = _parse_json_fields(dict(arow))
        return {"status": "ok", "news": news, "analysis": analysis}
    finally:
        cur.close()


def analyze_news(body):
    """Full analysis of a single news: viral, freshness, quality, relevance, sentiment, tags, trends, keyso."""
    news_id = body.get("news_id")
    if not news_id:
        return {"status": "error", "message": "news_id required"}
    conn = get_connection()
    cur = conn.cursor()
    try:
        ph = "%s" if _is_postgres() else "?"
        cur.execute(f"SELECT * FROM news WHERE id = {ph}", (news_id,))
        row = cur.fetchone()
        if not row:
            return {"status": "error", "message": "Not found"}
        if _is_postgres():
            columns = [desc[0] for desc in cur.description]
            news = dict(zip(columns, row))
        else:
            news = dict(row)

        from checks.viral_score import viral_score
        from checks.freshness import check_freshness
        from checks.quality import check_quality
        from checks.relevance import check_relevance
        from checks.sentiment import analyze_sentiment
        from checks.tags import auto_tag
        from checks.momentum import get_momentum

        result = {
            "viral": viral_score(news),
            "freshness": check_freshness(news),
            "quality": check_quality(news),
            "relevance": check_relevance(news),
            "sentiment": analyze_sentiment(news),
            "tags": auto_tag(news),
            "momentum": get_momentum(news),
        }

        cur.execute(f"SELECT * FROM news_analysis WHERE news_id = {ph}", (news_id,))
        arow = cur.fetchone()
        if arow:
            if _is_postgres():
                columns = [desc[0] for desc in cur.description]
                analysis = dict(zip(columns, arow))
            else:
                analysis = dict(arow)
            try:
                result["trends_data"] = json.loads(analysis.get("trends_data", "{}"))
            except Exception:
                result["trends_data"] = {}
            try:
                result["keyso_data"] = json.loads(analysis.get("keyso_data", "{}"))
            except Exception:
                result["keyso_data"] = {}
            result["llm_recommendation"] = analysis.get("llm_recommendation", "")
            result["llm_trend_forecast"] = analysis.get("llm_trend_forecast", "")
            try:
                result["bigrams"] = json.loads(analysis.get("bigrams", "[]"))
            except Exception:
                result["bigrams"] = []
        else:
            result["trends_data"] = {}
            result["keyso_data"] = {}
            result["llm_recommendation"] = ""
            result["llm_trend_forecast"] = ""
            result["bigrams"] = []

        total = sum(result[k]["score"] for k in ("viral", "freshness", "quality", "relevance")) // 4
        result["total_score"] = min(100, total + result["momentum"]["score"] // 5)
        return {"status": "ok", "analysis": result}
    finally:
        cur.close()


def merge_news(body):
    """Merge multiple news using LLM."""
    news_ids = body.get("news_ids", [])
    if len(news_ids) < 2:
        return {"status": "error", "message": "Need at least 2 news to merge"}
    try:
        conn = get_connection()
        cur = conn.cursor()
        try:
            ph = "%s" if _is_postgres() else "?"
            placeholders = ",".join([ph] * len(news_ids))
            cur.execute(f"SELECT id, source, title, plain_text FROM news WHERE id IN ({placeholders})", news_ids)
            if _is_postgres():
                columns = [desc[0] for desc in cur.description]
                news_list = [dict(zip(columns, row)) for row in cur.fetchall()]
            else:
                news_list = [dict(row) for row in cur.fetchall()]
            from apis.llm import merge_news as llm_merge_news
            result = llm_merge_news(news_list)
            if result:
                return {"status": "ok", "result": result, "sources": [n["source"] for n in news_list]}
            else:
                return {"status": "error", "message": "LLM returned no result"}
        finally:
            cur.close()
    except Exception as e:
        return {"status": "error", "message": str(e)}


# ---------------------------------------------------------------------------
# Sheets export
# ---------------------------------------------------------------------------

def export_sheets(body):
    """Export single news using current Sheets routing policy."""
    news_id = body.get("news_id")
    try:
        from storage.sheets import get_sheets_config_error
        from api.sheets_export import export_news_by_policy
        config_error = get_sheets_config_error()
        if config_error:
            return {"status": "error", "message": config_error}
        conn = get_connection()
        cur = conn.cursor()
        try:
            ph = "%s" if _is_postgres() else "?"
            cur.execute(f"SELECT * FROM news WHERE id = {ph}", (news_id,))
            if _is_postgres():
                columns = [desc[0] for desc in cur.description]
                news = dict(zip(columns, cur.fetchone()))
            else:
                news = dict(cur.fetchone())

            cur.execute(f"SELECT * FROM news_analysis WHERE news_id = {ph}", (news_id,))
            row = cur.fetchone()
            if row:
                if _is_postgres():
                    columns = [desc[0] for desc in cur.description]
                    analysis = dict(zip(columns, row))
                else:
                    analysis = dict(row)
            else:
                analysis = {"bigrams": "[]", "trends_data": "{}", "keyso_data": "{}",
                           "llm_recommendation": "", "llm_trend_forecast": "", "llm_merged_with": ""}

            result = export_news_by_policy(cur, news, analysis)
            row = result.get("row")
            if row and row > 0:
                return {"status": "ok", "row": row, "destination": result.get("destination")}
            if row == -1:
                return {"status": "ok", "row": -1, "destination": result.get("destination")}
            return {"status": "skipped", "destination": result.get("destination"), "message": result.get("reason", "Not eligible for export")}
        finally:
            cur.close()
    except Exception as e:
        return {"status": "error", "message": str(e)}


def export_sheets_bulk(body):
    """Export multiple news using current Sheets routing policy."""
    news_ids = body.get("news_ids", [])
    if not news_ids:
        return {"status": "error", "message": "news_ids required"}
    try:
        from storage.sheets import get_sheets_config_error
        from api.sheets_export import export_news_by_policy
        config_error = get_sheets_config_error()
        if config_error:
            return {"status": "error", "message": config_error}
        conn = get_connection()
        cur = conn.cursor()
        try:
            ph = "%s" if _is_postgres() else "?"
            exported = 0
            skipped = 0
            errors = 0
            ready = 0
            not_ready = 0
            for nid in news_ids:
                try:
                    cur.execute(f"SELECT * FROM news WHERE id = {ph}", (nid,))
                    row = cur.fetchone()
                    if not row:
                        continue
                    if _is_postgres():
                        columns = [desc[0] for desc in cur.description]
                        news = dict(zip(columns, row))
                    else:
                        news = dict(row)
                    cur.execute(f"SELECT * FROM news_analysis WHERE news_id = {ph}", (nid,))
                    arow = cur.fetchone()
                    if arow:
                        if _is_postgres():
                            columns = [desc[0] for desc in cur.description]
                            analysis = dict(zip(columns, arow))
                        else:
                            analysis = dict(arow)
                    else:
                        analysis = {"bigrams": "[]", "trends_data": "{}", "keyso_data": "{}",
                                   "llm_recommendation": "", "llm_trend_forecast": "", "llm_merged_with": ""}
                    result = export_news_by_policy(cur, news, analysis)
                    sheet_row = result.get("row")
                    if sheet_row and sheet_row > 0:
                        exported += 1
                        if result.get("destination") == "Ready":
                            ready += 1
                        elif result.get("destination") == "NotReady":
                            not_ready += 1
                    elif sheet_row == -1:
                        skipped += 1
                    else:
                        skipped += 1
                except Exception as e:
                    logger.warning("Bulk export error for %s: %s", nid, e)
                    errors += 1
            return {"status": "ok", "exported": exported, "skipped": skipped, "errors": errors, "ready": ready, "not_ready": not_ready}
        finally:
            cur.close()
    except Exception as e:
        return {"status": "error", "message": str(e)}


def export_all_processed(body):
    """Export all processed news to Sheets (background)."""
    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute("SELECT id FROM news WHERE status IN ('processed', 'ready') ORDER BY parsed_at DESC")
        if _is_postgres():
            news_ids = [r[0] for r in cur.fetchall()]
        else:
            news_ids = [r["id"] for r in cur.fetchall()]
    finally:
        cur.close()

    if not news_ids:
        return {"status": "error", "message": "Нет обработанных новостей"}

    from scheduler import _create_task
    task_ids = []
    cur2 = conn.cursor()
    ph = "%s" if _is_postgres() else "?"
    try:
        for nid in news_ids:
            try:
                cur2.execute(f"SELECT title FROM news WHERE id = {ph}", (nid,))
                row = cur2.fetchone()
                title = (row[0] if _is_postgres() else row["title"]) if row else ""
            except Exception:
                title = ""
            tid = _create_task("sheets", nid, title)
            task_ids.append(tid)
    finally:
        cur2.close()

    def _bg_export(ids, tids):
        from api.sheets_export import export_news_by_policy
        from scheduler import _update_task, _fetch_news_by_id, _fetch_analysis_by_id
        ok_count = 0
        skip_count = 0
        err_count = 0
        for i, (nid, tid) in enumerate(zip(ids, tids)):
            try:
                _update_task(tid, "running", {"stage": "exporting", "progress": f"{i+1}/{len(ids)}"})
                news = _fetch_news_by_id(nid)
                analysis = _fetch_analysis_by_id(nid)
                if not news:
                    _update_task(tid, "error", {"error": "News not found"})
                    err_count += 1
                    continue
                cur_local = get_connection().cursor()
                try:
                    result = export_news_by_policy(cur_local, news, analysis or {})
                finally:
                    cur_local.close()
                sheet_row = result.get("row")
                if sheet_row and sheet_row > 0:
                    _update_task(tid, "done", {"sheet_row": sheet_row, "destination": result.get("destination")})
                    ok_count += 1
                elif sheet_row == -1:
                    _update_task(tid, "skipped", {"reason": "duplicate in Sheets", "destination": result.get("destination")})
                    skip_count += 1
                else:
                    _update_task(tid, "skipped", {"reason": result.get("reason", "Not eligible for export")})
                    skip_count += 1
            except Exception as e:
                _update_task(tid, "error", {"error": str(e)[:500]})
                err_count += 1
            _time.sleep(1.5)
        logger.info("Mass export done: %d ok, %d skipped, %d errors out of %d", ok_count, skip_count, err_count, len(ids))

    threading.Thread(target=_bg_export, args=(list(news_ids), list(task_ids)), daemon=True).start()
    return {"status": "ok", "queued": len(news_ids), "task_ids": task_ids}


def export_ready_all(body):
    """Export ALL rewritten articles to Sheets/Ready (background)."""
    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT a.id, a.news_id, a.title, a.text, a.seo_title, a.seo_description,
                   a.tags, a.style, a.original_title, a.source_url, a.created_at,
                   n.source, n.parsed_at, n.url, n.title as news_title
            FROM articles a
            LEFT JOIN news n ON n.id = a.news_id
            ORDER BY a.created_at DESC
        """)
        if _is_postgres():
            columns = [d[0] for d in cur.description]
            articles = [dict(zip(columns, r)) for r in cur.fetchall()]
        else:
            articles = [dict(r) for r in cur.fetchall()]
    finally:
        cur.close()

    if not articles:
        return {"status": "error", "message": "Нет переписанных статей"}

    def _bg_export_ready(arts):
        from storage.sheets import write_ready_row
        from scheduler import _fetch_analysis_by_id
        import json as _json
        ok = 0
        skip = 0
        err = 0
        for art in arts:
            try:
                news_id = art.get("news_id", "")
                analysis = _fetch_analysis_by_id(news_id) if news_id else None

                news = {
                    "parsed_at": art.get("parsed_at", art.get("created_at", "")),
                    "source": art.get("source", ""),
                    "title": art.get("original_title") or art.get("news_title", ""),
                    "url": art.get("source_url") or art.get("url", ""),
                }

                tags_raw = art.get("tags", "[]")
                try:
                    tags_list = _json.loads(tags_raw) if isinstance(tags_raw, str) else (tags_raw if isinstance(tags_raw, list) else [])
                except Exception:
                    tags_list = []

                rewrite = {
                    "title": art.get("title", ""),
                    "text": art.get("text", ""),
                    "seo_title": art.get("seo_title", ""),
                    "seo_description": art.get("seo_description", ""),
                    "tags": tags_list,
                }

                row = write_ready_row(news, analysis, rewrite)
                if row and row > 0:
                    ok += 1
                elif row == -1:
                    skip += 1
                else:
                    err += 1
            except Exception as e:
                logger.error("Ready export error for article %s: %s", art.get("id"), e)
                err += 1
            _time.sleep(1.5)
        logger.info("Ready export done: %d ok, %d skipped, %d errors out of %d", ok, skip, err, len(arts))

    threading.Thread(target=_bg_export_ready, args=(list(articles),), daemon=True).start()
    return {"status": "ok", "queued": len(articles), "message": f"Экспорт {len(articles)} статей в Ready запущен"}


# ---------------------------------------------------------------------------
# SEO, moderation rewrite, rescore, quick tags, dashboard groups, translate, AI
# ---------------------------------------------------------------------------

def seo_check(body):
    """SEO analysis of an article."""
    from checks.seo_check import analyze_seo
    title = body.get("title", "")
    seo_title = body.get("seo_title", "")
    seo_description = body.get("seo_description", "")
    text = body.get("text", "")
    tags = body.get("tags", [])
    result = analyze_seo(title, seo_title, seo_description, text, tags)
    return {"status": "ok", **result}


def moderation_rewrite(body):
    """Send moderation news for rewrite (LLM only, no API enrichment)."""
    news_ids = body.get("news_ids", [])
    style = body.get("style", "news")
    if not news_ids:
        return {"status": "error", "message": "news_ids required"}

    from scheduler import _create_task
    task_ids = []
    conn = get_connection()
    cur = conn.cursor()
    try:
        ph = "%s" if _is_postgres() else "?"
        for nid in news_ids:
            try:
                cur.execute(f"SELECT title FROM news WHERE id = {ph}", (nid,))
                row = cur.fetchone()
                title = (row[0] if _is_postgres() else row["title"]) if row else ""
            except Exception:
                title = ""
            tid = _create_task("mod_rewrite", nid, title, style)
            task_ids.append(tid)
    finally:
        cur.close()

    def _bg_mod_rewrite(ids, tids, rewrite_style):
        from apis.llm import rewrite_news
        from scheduler import _update_task, _fetch_news_by_id, _fetch_analysis_by_id
        from storage.sheets import write_ready_row
        import uuid as _uuid
        import json as _json2
        for nid, tid in zip(ids, tids):
            try:
                news = _fetch_news_by_id(nid)
                if not news:
                    _update_task(tid, "error", {"error": "News not found"})
                    continue
                _update_task(tid, "running", {"stage": "rewriting"})
                result = rewrite_news(
                    title=news.get("title", ""),
                    text=news.get("plain_text", ""),
                    style=rewrite_style,
                    language="русский",
                )
                if result:
                    conn2 = get_connection()
                    cur2 = conn2.cursor()
                    ph2 = "%s" if _is_postgres() else "?"
                    _now = __import__('datetime').datetime.now(__import__('datetime').timezone.utc).isoformat()
                    aid = str(_uuid.uuid4())[:12]
                    tags_j = _json2.dumps(result.get("tags", []), ensure_ascii=False)
                    try:
                        cur2.execute(f"""INSERT INTO articles (id, news_id, title, text, seo_title, seo_description, tags,
                            style, language, original_title, original_text, source_url, status, created_at, updated_at)
                            VALUES ({','.join([ph2]*15)})""",
                            (aid, nid, result.get("title", ""), result.get("text", ""),
                             result.get("seo_title", ""), result.get("seo_description", ""), tags_j,
                             rewrite_style, "русский", news.get("title", ""), (news.get("plain_text", "") or "")[:5000],
                             news.get("url", ""), "draft", _now, _now))
                        if not _is_postgres():
                            conn2.commit()
                    finally:
                        cur2.close()

                    try:
                        analysis = _fetch_analysis_by_id(nid)
                        write_ready_row(news, analysis, result)
                    except Exception as se:
                        logger.warning("Sheets Ready export failed for %s: %s", nid, se)

                    _update_task(tid, "done", {
                        "stage": "complete",
                        "rewrite_title": result.get("title", "")[:100],
                        "article_id": aid,
                    })
                    update_news_status(nid, "processed")
                else:
                    _update_task(tid, "error", {"stage": "rewriting", "error": "Rewrite returned None"})
            except Exception as e:
                _update_task(tid, "error", {"error": str(e)[:500]})

    threading.Thread(
        target=_bg_mod_rewrite,
        args=(list(news_ids), list(task_ids), style),
        daemon=True
    ).start()

    return {"status": "ok", "queued": len(news_ids), "task_ids": task_ids}


def rescore_news(body):
    """Re-run scoring pipeline for specific news or all with score=0.
    Caller must check permissions before calling."""
    news_ids = body.get("news_ids", [])
    rescore_zero = body.get("rescore_zero", False)
    conn = get_connection()
    cur = conn.cursor()
    ph = "%s" if _is_postgres() else "?"
    try:
        if rescore_zero:
            cur.execute(f"""
                SELECT n.* FROM news n
                LEFT JOIN news_analysis a ON n.id = a.news_id
                WHERE n.status IN ('in_review', 'new', 'rejected')
                AND (a.total_score IS NULL OR a.total_score = 0 OR a.news_id IS NULL)
                ORDER BY n.parsed_at DESC
                LIMIT 500
            """)
        elif news_ids:
            placeholders = ",".join([ph] * len(news_ids))
            cur.execute(f"SELECT * FROM news WHERE id IN ({placeholders})", tuple(news_ids))
        else:
            return {"status": "error", "message": "news_ids or rescore_zero required"}

        if _is_postgres():
            columns = [desc[0] for desc in cur.description]
            news_list = [dict(zip(columns, r)) for r in cur.fetchall()]
        else:
            news_list = [dict(r) for r in cur.fetchall()]
    finally:
        cur.close()

    if not news_list:
        return {"status": "ok", "rescored": 0, "message": "Нет новостей для пересчёта"}

    from checks.pipeline import run_review_pipeline
    result = run_review_pipeline(news_list, update_status=False)
    scored = len(result.get("results", []))
    return {"status": "ok", "rescored": scored}


def quick_tags(body):
    """Quick tag calculation by headlines (without full review)."""
    news_ids = body.get("news_ids", [])
    if not news_ids:
        return {"status": "error", "message": "No news_ids"}
    try:
        conn = get_connection()
        cur = conn.cursor()
        try:
            ph = "%s" if _is_postgres() else "?"
            placeholders = ",".join([ph] * len(news_ids))
            cur.execute(f"SELECT id, title, description FROM news WHERE id IN ({placeholders})", news_ids)
            if _is_postgres():
                columns = [desc[0] for desc in cur.description]
                rows = [dict(zip(columns, row)) for row in cur.fetchall()]
            else:
                rows = [dict(row) for row in cur.fetchall()]

            from checks.tags import auto_tag
            from checks.deduplication import tfidf_similarity

            tags_map = {}
            for r in rows:
                tags = auto_tag(r)
                tags_map[r["id"]] = [{"id": t["id"], "label": t["label"], "hits": t["hits"]} for t in tags[:3]]

            titles = [r.get("title", "") for r in rows]
            ids_ordered = [r["id"] for r in rows]
            pairs = tfidf_similarity(titles)

            from collections import defaultdict
            graph = defaultdict(set)
            for i, j, score in pairs:
                graph[i].add(j)
                graph[j].add(i)
            visited = set()
            groups = []
            group_idx = 0
            id_to_group = {}
            for idx in range(len(rows)):
                if idx in visited:
                    continue
                cluster = set()
                stack = [idx]
                while stack:
                    node = stack.pop()
                    if node in visited:
                        continue
                    visited.add(node)
                    cluster.add(node)
                    stack.extend(graph[node] - visited)
                if len(cluster) >= 2:
                    group_idx += 1
                    member_ids = [ids_ordered[i] for i in sorted(cluster)]
                    member_titles = [titles[i] for i in sorted(cluster)]
                    for mid in member_ids:
                        id_to_group[mid] = group_idx
                    groups.append({
                        "group": group_idx,
                        "count": len(member_ids),
                        "ids": member_ids,
                        "titles": member_titles,
                    })

            return {"status": "ok", "tags": tags_map, "groups": groups, "id_to_group": id_to_group}
        finally:
            cur.close()
    except Exception as e:
        return {"status": "error", "message": str(e), "type": type(e).__name__}


def dashboard_groups(query_params):
    """Returns tags and groups for news (respects status filter)."""
    try:
        conn = get_connection()
        cur = conn.cursor()
        try:
            ph = "%s" if _is_postgres() else "?"
            status_filter = query_params.get("status", [None])[0]
            if status_filter:
                cur.execute(f"SELECT id, title, description FROM news WHERE status = {ph} ORDER BY parsed_at DESC LIMIT 100", (status_filter,))
            else:
                cur.execute("SELECT id, title, description FROM news ORDER BY parsed_at DESC LIMIT 100")
            if _is_postgres():
                columns = [desc[0] for desc in cur.description]
                rows = [dict(zip(columns, row)) for row in cur.fetchall()]
            else:
                rows = [dict(row) for row in cur.fetchall()]

            if not rows:
                return {"status": "ok", "tags": {}, "groups": [], "id_to_group": {}}

            from checks.tags import auto_tag
            from checks.deduplication import tfidf_similarity
            from collections import defaultdict

            tags_map = {}
            for r in rows:
                tags = auto_tag(r)
                tags_map[r["id"]] = [{"id": t["id"], "label": t["label"], "hits": t["hits"]} for t in tags[:3]]

            titles = [r.get("title", "") for r in rows]
            ids_ordered = [r["id"] for r in rows]
            pairs = tfidf_similarity(titles)

            graph = defaultdict(set)
            for i, j, score in pairs:
                graph[i].add(j)
                graph[j].add(i)
            visited = set()
            groups = []
            group_idx = 0
            id_to_group = {}
            for idx in range(len(rows)):
                if idx in visited:
                    continue
                cluster = set()
                stack = [idx]
                while stack:
                    node = stack.pop()
                    if node in visited:
                        continue
                    visited.add(node)
                    cluster.add(node)
                    stack.extend(graph[node] - visited)
                if len(cluster) >= 2:
                    group_idx += 1
                    member_ids = [ids_ordered[i] for i in sorted(cluster)]
                    member_titles = [titles[i] for i in sorted(cluster)]
                    for mid in member_ids:
                        id_to_group[mid] = group_idx
                    groups.append({
                        "group": group_idx,
                        "count": len(member_ids),
                        "ids": member_ids,
                        "titles": member_titles,
                    })

            return {"status": "ok", "tags": tags_map, "groups": groups, "id_to_group": id_to_group}
        finally:
            cur.close()
    except Exception as e:
        return {"status": "error", "message": str(e), "type": type(e).__name__}


def translate_title(body):
    """Translate news title using LLM."""
    news_id = body.get("news_id", "")
    if not news_id:
        return {"status": "error", "message": "news_id required"}
    conn = get_connection()
    cur = conn.cursor()
    try:
        ph = "%s" if _is_postgres() else "?"
        cur.execute(f"SELECT title FROM news WHERE id = {ph}", (news_id,))
        row = cur.fetchone()
        if not row:
            return {"status": "error", "message": "not found"}
        title = row[0] if _is_postgres() else row["title"]
        try:
            from apis.llm import translate_title as llm_translate_title
            result = llm_translate_title(title)
            if result:
                if not result.get("is_russian") and result.get("translated"):
                    cur.execute(f"UPDATE news SET h1 = {ph} WHERE id = {ph}", (result["translated"], news_id))
                    if not _is_postgres():
                        conn.commit()
                return {"status": "ok", **result}
            else:
                return {"status": "error", "message": "LLM not responding. Check API keys and rate limits."}
        except Exception as e:
            logger.error("Translate error: %s", e)
            return {"status": "error", "message": str(e)}
    finally:
        cur.close()


def ai_recommend(body):
    """Get AI recommendation for a news item."""
    news_id = body.get("news_id", "")
    if not news_id:
        return {"status": "error", "message": "news_id required"}
    conn = get_connection()
    cur = conn.cursor()
    try:
        ph = "%s" if _is_postgres() else "?"
        cur.execute(f"SELECT * FROM news WHERE id = {ph}", (news_id,))
        row = cur.fetchone()
        if not row:
            return {"status": "error", "message": "not found"}
        if _is_postgres():
            columns = [desc[0] for desc in cur.description]
            news = dict(zip(columns, row))
        else:
            news = dict(row)

        from checks.quality import check_quality
        from checks.relevance import check_relevance
        from checks.freshness import check_freshness
        from checks.viral_score import viral_score
        checks = {
            "quality": check_quality(news),
            "relevance": check_relevance(news),
            "freshness": check_freshness(news),
            "viral": viral_score(news),
        }
        from apis.llm import ai_recommendation
        result = ai_recommendation(
            title=news.get("title", ""),
            text=news.get("plain_text", "") or news.get("description", ""),
            source=news.get("source", ""),
            checks=checks,
        )
        if result:
            return {"status": "ok", "recommendation": result, "checks": {k: v.get("score", 0) for k, v in checks.items()}}
        else:
            return {"status": "error", "message": "AI recommendation failed"}
    finally:
        cur.close()
